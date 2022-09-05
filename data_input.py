import numpy
import numpy as np
import pandas as pd
import requests
import math
import openpyxl as xl
import time
import datetime
import os
import glob


stock_list = []
pd.set_option('display.max_rows', 500)

#Removes old files
path = 'C:\\Users\\aidan\\PycharmProjects\\New Algo\\'
for f in glob.iglob(path + '/**/*.xlsx', recursive=True):
    os.remove(f)


# Ask user for stock and add to stock_list
def input_stocks():
    global stock_list
    user_input = input('Enter your tickers separated by a "space" please.')
    stock_list = user_input.split()


#grab the data from yahoo finance and download
def collect_data():
    for received_ticker in stock_list:
        ticker = str(received_ticker)
        period1 = int(time.mktime(datetime.datetime(2022, 1, 1, 23, 59).timetuple()))
        period2 = int(time.mktime(datetime.datetime(2022, 8, 30, 23, 59).timetuple()))
        interval = '1d'
        query_string = f'https://query1.finance.yahoo.com/v7/finance/download/{ticker}?period1={period1}&period2={period2}&interval={interval}&events=history&includeAdjustedClose=true'
        df = pd.read_csv(query_string)
        df = df.drop('Open', axis=1)
        df = df.drop('High', axis=1)
        df = df.drop('Low', axis=1)
        df = df.drop('Close', axis=1)
        df = df.drop('Volume', axis=1)
        writer = pd.ExcelWriter(f'{ticker}-to-excel.xlsx')
        df.to_excel(writer, index=False)
        writer.save()


# Combines our excel files collected in main folder to combined data folder puts in individual sheets
def combine_excel_files():
    location = 'C:\\Users\\aidan\\PycharmProjects\\New Algo\\*.xlsx'
    excel_files = glob.glob(location)
    writer2 = pd.ExcelWriter('C:\\Users\\aidan\\PycharmProjects\\New Algo\\Combined Data\\Combined_excel.xlsx')

    for excel_file in excel_files:
        sheet = os.path.basename(excel_file)
        sheet = sheet.split('.')[0]
        df1 = pd.read_excel(excel_file)
        df1.fillna(value='N/A', inplace=True)
        df1.to_excel(writer2, sheet_name=sheet, index=False)

    writer2.save()

# copy data from stock sheets to combined sheet[0]
    wb1 = xl.load_workbook('C:\\Users\\aidan\\PycharmProjects\\New Algo\\Combined Data\\Combined_excel.xlsx')
    wb1.create_sheet('All data', 0)
    for index, stock in enumerate(stock_list, start=1):
        master_sheet = wb1.worksheets[0]
        ws2 = wb1.worksheets[index]
        max_rows = ws2.max_row
        max_columns = ws2.max_column
        position = index
        for i in range(1, max_rows + 1):
            for j in range(1, max_columns + 1):
                c = ws2.cell(row=i, column=j)
                master_sheet.cell(row=i, column=position).value = c.value
        position += 2

# Insert Date Column
    master_sheet.insert_cols(0)
    for i in range(1, max_rows + 1):
        for j in range(1, 2):
            c = ws2.cell(row=i, column=j)
            master_sheet.cell(row=i, column=1).value = c.value

# Insert Row with Stock Names
    master_sheet.insert_rows(0)
    for index, stock in enumerate(stock_list, start=1):
        max_columns = ws2.max_column
        for j in range(1, max_columns):
            master_sheet.cell(row=1, column=j+index).value = str(stock)
    master_sheet.cell(row=1, column=1).value = 'Date'
    master_sheet.delete_rows(2)

    wb1.save('C:\\Users\\aidan\\PycharmProjects\\New Algo\\Combined Data\\Combined_excel.xlsx')


input_stocks()
collect_data()
combine_excel_files()